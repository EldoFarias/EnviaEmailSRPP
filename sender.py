#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Envio Automático de PDFs por Email
Monitora pasta de PDFs e processa automaticamente quando novos arquivos são criados
"""

import os
import re
import glob
import smtplib
import logging
import pyodbc
import time
import signal
import sys
import hashlib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

# Para logs em Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_DISPONIVEL = True
except ImportError:
    EXCEL_DISPONIVEL = False

# Para monitoramento de arquivos
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

class SistemaEnvioEmails:
    def __init__(self):
        """Inicializa o sistema com configurações"""
        # Configurações de conexão SQL embutidas (não mudam nunca)
        self.sql_config = {
            'servidor': '127.0.0.1',
            'banco_de_dados': 'SRPP',
            'usuario': 'sa',
            'senha': 'M4573R',
            'driver_preferencial': 'ODBC Driver 17 for SQL Server'
        }
        self.config_db = None  # Configurações carregadas do banco de dados
        self.setup_logging()
        self.conexao_db = None
        self.excel_logger = None
        self.carregar_configuracoes_banco()  # Tenta carregar configurações do banco
        self.setup_excel_logging()

    def carregar_configuracoes_banco(self):
        """Carrega configurações da tabela ConfiguracaoSistemaEmail no banco de dados"""
        try:
            # Conecta ao banco usando credenciais embutidas
            if not self.conectar_banco():
                self.logger.error("ERRO CRÍTICO: Não foi possível conectar ao banco de dados. Verifique se o SQL Server está acessível.")
                return False

            cursor = self.conexao_db.cursor()
            query = """
                SELECT TOP 1
                    SqlServidor, SqlBancoDados, SqlUsuario, SqlSenha, SqlDriver,
                    PdfsCaminho,
                    EmailSmtpServidor, EmailSmtpPorta, EmailUsuario, EmailSenhaApp, EmailRemetente,
                    EmailAssunto, EmailCorpo, EmailResponderPara,
                    SistemaVerificacaoInicial, SistemaAguardarSegundosAposArquivo,
                    SistemaVerificacaoPeriodicaAtiva, SistemaVerificacaoPeriodicaMinutos,
                    SistemaCooldownTentativa1, SistemaCooldownTentativa2, SistemaCooldownTentativa3,
                    SistemaCooldownTentativa4, SistemaCooldownTentativa5Mais
                FROM ConfiguracaoSistemaEmail
                WHERE Ativo = 1
            """
            cursor.execute(query)
            row = cursor.fetchone()

            if row:
                self.config_db = {
                    'sql_servidor': row.SqlServidor,
                    'sql_banco_dados': row.SqlBancoDados,
                    'sql_usuario': row.SqlUsuario,
                    'sql_senha': row.SqlSenha,
                    'sql_driver': row.SqlDriver,
                    'pdfs_caminho': row.PdfsCaminho,
                    'email_smtp_servidor': row.EmailSmtpServidor,
                    'email_smtp_porta': row.EmailSmtpPorta,
                    'email_usuario': row.EmailUsuario,
                    'email_senha_app': row.EmailSenhaApp,
                    'email_remetente': row.EmailRemetente,
                    'email_assunto': row.EmailAssunto,
                    'email_corpo': row.EmailCorpo,
                    'email_responder_para': row.EmailResponderPara,
                    'sistema_verificacao_inicial': row.SistemaVerificacaoInicial,
                    'sistema_aguardar_segundos': row.SistemaAguardarSegundosAposArquivo,
                    'sistema_verificacao_periodica_ativa': row.SistemaVerificacaoPeriodicaAtiva,
                    'sistema_verificacao_periodica_minutos': row.SistemaVerificacaoPeriodicaMinutos,
                    'cooldown_tentativa_1': row.SistemaCooldownTentativa1,
                    'cooldown_tentativa_2': row.SistemaCooldownTentativa2,
                    'cooldown_tentativa_3': row.SistemaCooldownTentativa3,
                    'cooldown_tentativa_4': row.SistemaCooldownTentativa4,
                    'cooldown_tentativa_5_mais': row.SistemaCooldownTentativa5Mais,
                }
                self.logger.info("Configurações carregadas do banco de dados com sucesso!")
                self.conexao_db.close()
                self.conexao_db = None
                return True
            else:
                self.logger.error("ERRO CRÍTICO: Nenhuma configuração ativa encontrada na tabela ConfiguracaoSistemaEmail.")
                self.logger.error("SOLUÇÃO: Execute o script SQL 'criar_tabela_configuracoes.sql' para criar a tabela.")
                self.conexao_db.close()
                self.conexao_db = None
                return False

        except Exception as e:
            self.logger.error(f"ERRO ao carregar configurações do banco: {e}")
            if self.conexao_db:
                self.conexao_db.close()
                self.conexao_db = None
            return False

    def get_config(self, secao, chave, fallback=None):
        """Obtém configuração do banco de dados"""
        if not self.config_db:
            self.logger.error(f"ERRO: Tentativa de obter configuração '{secao}.{chave}' mas config_db não está carregado!")
            return fallback

        # Mapeia seção/chave para as chaves do dicionário do banco
        mapa = {
            ('PDFS', 'caminho'): 'pdfs_caminho',
            ('EMAIL', 'smtp_servidor'): 'email_smtp_servidor',
            ('EMAIL', 'smtp_porta'): 'email_smtp_porta',
            ('EMAIL', 'usuario'): 'email_usuario',
            ('EMAIL', 'senha_app'): 'email_senha_app',
            ('EMAIL', 'remetente_nome'): 'email_remetente',
            ('EMAIL', 'reply_to'): 'email_responder_para',
            ('SISTEMA', 'verificacao_inicial'): 'sistema_verificacao_inicial',
            ('SISTEMA', 'aguardar_segundos_apos_arquivo'): 'sistema_aguardar_segundos',
            ('SISTEMA', 'verificacao_periodica_ativa'): 'sistema_verificacao_periodica_ativa',
            ('SISTEMA', 'verificacao_periodica_minutos'): 'sistema_verificacao_periodica_minutos',
            ('SISTEMA', 'cooldown_tentativa_1'): 'cooldown_tentativa_1',
            ('SISTEMA', 'cooldown_tentativa_2'): 'cooldown_tentativa_2',
            ('SISTEMA', 'cooldown_tentativa_3'): 'cooldown_tentativa_3',
            ('SISTEMA', 'cooldown_tentativa_4'): 'cooldown_tentativa_4',
            ('SISTEMA', 'cooldown_tentativa_5_mais'): 'cooldown_tentativa_5_mais',
        }

        chave_db = mapa.get((secao, chave))
        if chave_db and chave_db in self.config_db:
            return self.config_db[chave_db]

        return fallback
        
    def setup_logging(self):
        """Configura sistema de logs"""
        log_dir = 'logs'
        os.makedirs(log_dir, exist_ok=True)
        
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(f'{log_dir}/envio_emails.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
    def setup_excel_logging(self):
        """Configura sistema de logs em Excel"""
        if not EXCEL_DISPONIVEL:
            self.logger.warning("openpyxl não disponível - logs em Excel desabilitados")
            return

        try:
            self.excel_logger = ExcelLogger()
            self.logger.info("Sistema de logs Excel iniciado")
        except Exception as e:
            self.logger.error(f"Erro ao inicializar logs Excel: {e}")
            self.excel_logger = None

    def listar_drivers_odbc_disponiveis(self):
        """Lista todos os drivers ODBC instalados no sistema"""
        try:
            drivers = pyodbc.drivers()
            self.logger.info(f"Drivers ODBC instalados no sistema: {len(drivers)}")
            for driver in drivers:
                self.logger.info(f"  - {driver}")
            return drivers
        except Exception as e:
            self.logger.error(f"Erro ao listar drivers ODBC: {e}")
            return []
        
    def conectar_banco(self):
        """Estabelece conexão com SQL Server com fallback automático de drivers"""
        # Lista de drivers ODBC em ordem de preferência (mais recente para mais antigo)
        drivers_disponiveis = [
            self.sql_config['driver_preferencial'],  # Driver preferencial
            'ODBC Driver 18 for SQL Server',
            'ODBC Driver 17 for SQL Server',
            'ODBC Driver 13.1 for SQL Server',
            'ODBC Driver 13 for SQL Server',
            'ODBC Driver 11 for SQL Server',
            'SQL Server Native Client 11.0',
            'SQL Server Native Client 10.0',
            'SQL Server',
        ]

        # Remove duplicatas mantendo a ordem
        drivers_unicos = []
        for driver in drivers_disponiveis:
            if driver not in drivers_unicos:
                drivers_unicos.append(driver)

        erros_tentativas = []
        driver_preferencial = self.sql_config['driver_preferencial']

        for driver in drivers_unicos:
            try:
                conn_str = (
                    f"DRIVER={{{driver}}};"
                    f"SERVER={self.sql_config['servidor']};"
                    f"DATABASE={self.sql_config['banco_de_dados']};"
                    f"UID={self.sql_config['usuario']};"
                    f"PWD={self.sql_config['senha']};"
                )
                self.conexao_db = pyodbc.connect(conn_str)
                self.logger.info(f"Conectado ao banco de dados com sucesso usando driver: {driver}")

                # Se conectou com driver diferente do preferencial, avisar
                if driver != driver_preferencial:
                    self.logger.warning(f"Driver preferencial '{driver_preferencial}' não disponível. Usando '{driver}' como alternativa.")

                return True

            except Exception as e:
                erros_tentativas.append(f"{driver}: {str(e)[:100]}")
                # Continua tentando próximo driver

        # Se chegou aqui, nenhum driver funcionou
        self.logger.error("ERRO: Não foi possível conectar ao banco de dados com nenhum driver ODBC disponível")
        self.logger.error("Drivers tentados:")
        for erro in erros_tentativas:
            self.logger.error(f"  - {erro}")
        self.logger.error("SOLUÇÃO: Instale um driver ODBC para SQL Server:")
        self.logger.error("  - Download: https://go.microsoft.com/fwlink/?linkid=2249004")

        return False
            
    def buscar_pedidos_para_processar(self):
        """Busca todos os pedidos que precisam ser processados"""
        query = """
        SELECT
            cep.Id, cep.NroPedido, cep.CodCliente, cep.DataPedidoFechado, cep.EmailsCopia,
            c.EMAIL as EmailCliente, c.NomeContato as NomeCliente,
            cep.VersaoPdfEnviada, cep.StatusProcessamento, cep.EmailEnviado,
            cep.TentativasEnvio, NULL as DataUltimaVerificacao, cep.UltimoErro,
            cep.EnviarEmailCliente
        FROM ControleEmailPedidos cep
        INNER JOIN CabecalhoPedido cap ON cap.NroPedido = cep.NroPedido
        INNER JOIN Cliente c ON c.CodCliente = cap.CodCliente
        WHERE cap.SituacaoAtual = 'F'
            AND (
                (cep.StatusProcessamento = 'PENDENTE' AND cep.EmailEnviado = 0) OR
                (cep.StatusProcessamento = 'ENVIADO' AND cep.EmailEnviado = 1) OR
                (cep.StatusProcessamento = 'ERRO_VALIDACAO' AND cep.EmailEnviado = 0) OR
                (cep.StatusProcessamento = 'INVALIDO' AND cep.EmailEnviado = 0) OR
                (cep.StatusProcessamento = 'ERRO' AND cep.EmailEnviado = 0)
            )
        ORDER BY cep.DataPedidoFechado
        """
        
        try:
            cursor = self.conexao_db.cursor()
            cursor.execute(query)
            resultados = cursor.fetchall()
            
            pedidos_para_processar = []
            
            for row in resultados:
                numero_pedido = row.NroPedido
                versao_enviada = int(row.VersaoPdfEnviada) if row.VersaoPdfEnviada is not None else 0
                status_atual = row.StatusProcessamento

                caminho_pdf, versao_disponivel = self.buscar_pdf_pedido(numero_pedido)

                deve_processar = False
                motivo = ""

                if not caminho_pdf:
                    self.logger.warning(f"Pedido {numero_pedido}: PDF não encontrado, pulando...")
                    continue
                elif status_atual in ('PENDENTE', 'INVALIDO') and row.EmailEnviado == 0:
                    deve_processar = True
                    motivo = "PRIMEIRO_ENVIO"
                elif status_atual == 'ENVIADO' and versao_disponivel > versao_enviada:
                    deve_processar = True
                    motivo = "REENVIO_VERSAO_ATUALIZADA"
                elif status_atual == 'ERRO_VALIDACAO' and row.EmailEnviado == 0:
                    deve_processar = True
                    motivo = "REVALIDACAO_APOS_ERRO"
                
                if deve_processar:
                    pedidos_para_processar.append({
                        'id': row.Id, 'numero': numero_pedido, 'cod_cliente': row.CodCliente,
                        'data_fechamento': row.DataPedidoFechado, 'emails_copia': row.EmailsCopia,
                        'email_cliente': row.EmailCliente, 'nome_cliente': row.NomeCliente,
                        'versao_pdf_enviada': versao_enviada, 'status_atual': status_atual,
                        'versao_disponivel': versao_disponivel, 'motivo_processamento': motivo,
                        'caminho_pdf': caminho_pdf, 'tentativas_anteriores': row.TentativasEnvio or 0,
                        'ultimo_erro': row.UltimoErro, 'enviar_email_cliente': row.EnviarEmailCliente
                    })
                    
            self.logger.info(f"Encontrados {len(pedidos_para_processar)} pedidos para processar")
            return pedidos_para_processar
            
        except Exception as e:
            self.logger.error(f"Erro ao buscar pedidos para processar: {e}")
            return []
            
    def _calcular_cooldown(self, tentativas):
        """Calcula tempo de cool-down progressivo"""
        try:
            if tentativas <= 1: return int(self.get_config('SISTEMA', 'cooldown_tentativa_1', fallback=2))
            elif tentativas == 2: return int(self.get_config('SISTEMA', 'cooldown_tentativa_2', fallback=5))
            elif tentativas == 3: return int(self.get_config('SISTEMA', 'cooldown_tentativa_3', fallback=10))
            elif tentativas == 4: return int(self.get_config('SISTEMA', 'cooldown_tentativa_4', fallback=20))
            else: return int(self.get_config('SISTEMA', 'cooldown_tentativa_5_mais', fallback=30))
        except:
            if tentativas <= 1: return 2
            elif tentativas == 2: return 5
            elif tentativas == 3: return 10
            elif tentativas == 4: return 20
            else: return 30
        
    def buscar_pdf_pedido(self, numero_pedido):
        """Busca e identifica a versão mais recente do PDF"""
        caminho_pdfs = self.get_config('PDFS', 'caminho')
        padrao = f"PEDIDO {str(numero_pedido).zfill(7)}*.pdf"
        arquivos = glob.glob(os.path.join(caminho_pdfs, padrao))
        
        if not arquivos:
            self.logger.warning(f"Nenhum PDF encontrado para pedido {numero_pedido}")
            return None, 0
            
        versao_maxima = 1
        arquivo_mais_recente = None
        for arquivo in arquivos:
            match = re.search(r'PEDIDO \d+(?:_(\d+))?\.pdf$', os.path.basename(arquivo))
            if match:
                versao = int(match.group(1)) if match.group(1) else 1
                if versao >= versao_maxima:
                    versao_maxima = versao
                    arquivo_mais_recente = arquivo
        return arquivo_mais_recente, versao_maxima
        
    def atualizar_status_pedido(self, id_controle, campo, valor, erro=None):
        """Atualiza status do pedido na tabela de controle"""
        try:
            cursor = self.conexao_db.cursor()
            if erro:
                query = f"UPDATE ControleEmailPedidos SET {campo} = ?, UltimoErro = ?, TentativasEnvio = TentativasEnvio + 1 WHERE Id = ?"
                cursor.execute(query, (valor, str(erro)[:500], id_controle))
            else:
                query = f"UPDATE ControleEmailPedidos SET {campo} = ? WHERE Id = ?"
                cursor.execute(query, (valor, id_controle))
            self.conexao_db.commit()
        except Exception as e:
            self.logger.error(f"Erro ao atualizar status do pedido {id_controle}: {e}")
            
    def enviar_email(self, destinatario, nome_cliente, numero_pedido, caminho_pdf, emails_copia=None, eh_reenvio=False, versao_pdf=1, enviar_para_cliente=True, data_pedido_fechado=None):
        """Envia email com PDF anexo usando templates do banco de dados"""
        try:
            # Determinar destinatários baseado em enviar_para_cliente
            destinatario_principal = None
            lista_copia = []

            if enviar_para_cliente:
                # Fluxo normal: envia para cliente e cópias
                destinatario_principal = destinatario
                if emails_copia:
                    lista_copia = [email.strip() for email in emails_copia.split(',') if email.strip()]

                if lista_copia:
                    self.logger.info(f"Pedido {numero_pedido}: Enviando PARA cliente ({destinatario}) e CÓPIAS ({len(lista_copia)} email(s): {', '.join(lista_copia)})")
                else:
                    self.logger.info(f"Pedido {numero_pedido}: Enviando apenas PARA cliente ({destinatario}) - Sem emails em cópia")
            else:
                # Fluxo alternativo: envia apenas para emails de cópia
                if emails_copia:
                    lista_emails = [email.strip() for email in emails_copia.split(',') if email.strip()]
                    if lista_emails:
                        destinatario_principal = lista_emails[0]  # Primeiro vai para PARA
                        lista_copia = lista_emails[1:] if len(lista_emails) > 1 else []  # Demais vão para CÓPIAS

                        if lista_copia:
                            self.logger.info(f"Pedido {numero_pedido}: Cliente NÃO receberá email. Enviando PARA ({destinatario_principal}) e CÓPIAS ({len(lista_copia)} email(s): {', '.join(lista_copia)})")
                        else:
                            self.logger.info(f"Pedido {numero_pedido}: Cliente NÃO receberá email. Enviando apenas PARA ({destinatario_principal})")
                    else:
                        self.logger.warning(f"Pedido {numero_pedido}: EnviarEmailCliente=0 e EmailsCopia vazio - Nenhum destinatário definido")
                        return False
                else:
                    self.logger.warning(f"Pedido {numero_pedido}: EnviarEmailCliente=0 e EmailsCopia vazio - Nenhum destinatário definido")
                    return False

            # Obter templates do banco de dados ou usar padrões
            assunto_template = self.config_db.get('email_assunto') if self.config_db else 'Pedido {NroPedido} - PDF Anexado'
            corpo_template = self.config_db.get('email_corpo') if self.config_db else """Prezado(a) {RazaoSocial},

Segue em anexo o PDF do seu pedido número {NroPedido}.

{MensagemReenvio}

Atenciosamente,
Equipe SRPP"""

            # Detectar tipo de conteúdo (HTML ou TEXTO) e processar corpo
            tipo_conteudo = 'plain'  # Padrão é texto simples
            linhas = corpo_template.split('\n')

            if linhas and linhas[0].strip().upper() in ('#HTML', '#TEXTO'):
                tipo_marcador = linhas[0].strip().upper()
                # Remove a primeira linha (marcador) do template
                corpo_template = '\n'.join(linhas[1:])

                if tipo_marcador == '#HTML':
                    tipo_conteudo = 'html'
                    self.logger.debug(f"Pedido {numero_pedido}: Email será enviado como HTML")
                else:
                    self.logger.debug(f"Pedido {numero_pedido}: Email será enviado como TEXTO")

            # Preparar variáveis para substituição
            data_formatada = data_pedido_fechado.strftime("%d/%m/%Y") if data_pedido_fechado else datetime.now().strftime("%d/%m/%Y")

            # Mensagem de reenvio
            if eh_reenvio:
                if tipo_conteudo == 'html':
                    mensagem_reenvio = f"<strong>ATENÇÃO:</strong> Esta é uma versão atualizada do pedido (versão {versao_pdf}).<br>Esta versão substitui a anterior."
                else:
                    mensagem_reenvio = f"ATENÇÃO: Esta é uma versão atualizada do pedido (versão {versao_pdf}).\nEsta versão substitui a anterior."
            else:
                mensagem_reenvio = ""

            # Substituir variáveis no assunto
            assunto = assunto_template.format(
                NroPedido=numero_pedido,
                RazaoSocial=nome_cliente,
                DataPedidoFechado=data_formatada,
                VersaoPdf=versao_pdf,
                MensagemReenvio=mensagem_reenvio
            )

            # Substituir variáveis no corpo
            corpo = corpo_template.format(
                NroPedido=numero_pedido,
                RazaoSocial=nome_cliente,
                DataPedidoFechado=data_formatada,
                VersaoPdf=versao_pdf,
                MensagemReenvio=mensagem_reenvio
            )

            # Montar email
            msg = MIMEMultipart()
            email_usuario = self.get_config('EMAIL', 'usuario')  # Usado para autenticação SMTP
            email_remetente = self.get_config('EMAIL', 'remetente_nome')  # EmailRemetente do banco (FROM)
            email_reply_to = self.get_config('EMAIL', 'reply_to')  # EmailResponderPara (Reply-To)

            # Determinar o endereço FROM (quem "envia" o email)
            # Se EmailRemetente está configurado, usa ele (Send As com nao-responda@sint.com.br)
            # Caso contrário, usa email_usuario
            email_from = email_remetente if email_remetente else email_usuario

            # Montar o cabeçalho From (apenas o email, sem nome personalizado neste caso)
            msg['From'] = email_from
            msg['To'] = destinatario_principal
            msg['Subject'] = assunto

            # Configurar Reply-To se estiver definido no banco
            if email_reply_to:
                msg['Reply-To'] = email_reply_to
                self.logger.debug(f"Pedido {numero_pedido}: Autenticando como {email_usuario}, FROM={email_from}, Reply-To={email_reply_to}")
            else:
                self.logger.debug(f"Pedido {numero_pedido}: Autenticando como {email_usuario}, FROM={email_from}")

            if lista_copia:
                msg['Cc'] = ', '.join(lista_copia)

            msg.attach(MIMEText(corpo.strip(), tipo_conteudo, 'utf-8'))

            # Anexar PDF
            with open(caminho_pdf, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            nome_arquivo = f"Pedido_{numero_pedido}_v{versao_pdf}.pdf" if eh_reenvio else f"Pedido_{numero_pedido}.pdf"
            part.add_header('Content-Disposition', f'attachment; filename= {nome_arquivo}')
            msg.attach(part)

            # Enviar para todos os destinatários
            todos_destinatarios = [destinatario_principal] + lista_copia

            # Conectar ao servidor SMTP
            smtp_servidor = self.get_config('EMAIL', 'smtp_servidor')
            smtp_porta = int(self.get_config('EMAIL', 'smtp_porta', fallback=587))
            self.logger.debug(f"Pedido {numero_pedido}: Conectando ao servidor SMTP {smtp_servidor}:{smtp_porta}")
            server = smtplib.SMTP(smtp_servidor, smtp_porta)
            server.starttls()
            self.logger.debug(f"Pedido {numero_pedido}: Autenticando como {email_usuario}")
            senha_app = self.get_config('EMAIL', 'senha_app')
            server.login(email_usuario, senha_app)
            self.logger.debug(f"Pedido {numero_pedido}: Enviando email FROM {email_from} para {len(todos_destinatarios)} destinatário(s)")
            server.sendmail(email_from, todos_destinatarios, msg.as_string())
            server.quit()

            self.logger.info(f"{'REENVIO' if eh_reenvio else 'EMAIL'} enviado com sucesso - Pedido {numero_pedido} - Total de destinatários: {len(todos_destinatarios)}")
            return True
        except OSError as e:
            # Erros de rede (DNS, conexão, timeout)
            smtp_servidor = self.get_config('EMAIL', 'smtp_servidor', fallback='servidor_smtp')
            if 'getaddrinfo failed' in str(e) or '11001' in str(e):
                self.logger.error(f"Pedido {numero_pedido}: ERRO DE REDE - Falha ao resolver nome do servidor SMTP '{smtp_servidor}'. Verifique sua conexão com a internet.")
            else:
                self.logger.error(f"Pedido {numero_pedido}: ERRO DE REDE - {e}")
            return False
        except smtplib.SMTPAuthenticationError as e:
            self.logger.error(f"Pedido {numero_pedido}: ERRO DE AUTENTICAÇÃO - Verifique usuário/senha do email - {e}")
            return False
        except smtplib.SMTPException as e:
            self.logger.error(f"Pedido {numero_pedido}: ERRO SMTP - {e}")
            return False
        except Exception as e:
            self.logger.error(f"Pedido {numero_pedido}: ERRO INESPERADO ao enviar email - {e}")
            return False
            
    def processar_pedido(self, pedido):
        """Processa um pedido individual"""
        id_controle, numero_pedido = pedido['id'], pedido['numero']
        self.logger.info(f"Processando pedido {numero_pedido}...")
        self.atualizar_status_pedido(id_controle, 'StatusProcessamento', 'PROCESSANDO')

        try:
            if not self.validacao_geral(pedido): return False

            eh_reenvio = (pedido['motivo_processamento'] == "REENVIO_VERSAO_ATUALIZADA")
            enviar_para_cliente = pedido.get('enviar_email_cliente', 1) == 1  # Converte para boolean

            if self.enviar_email(
                pedido['email_cliente'],
                pedido['nome_cliente'],
                numero_pedido,
                pedido['caminho_pdf'],
                pedido['emails_copia'],
                eh_reenvio,
                pedido['versao_disponivel'],
                enviar_para_cliente,
                pedido['data_fechamento']
            ):
                query = "UPDATE ControleEmailPedidos SET StatusProcessamento = 'ENVIADO', EmailEnviado = 1, VersaoPdfEnviada = ?, DataEnvio = GETDATE(), UltimoErro = NULL, TentativasEnvio = 0 WHERE Id = ?"
                cursor = self.conexao_db.cursor()
                cursor.execute(query, (pedido['versao_disponivel'], id_controle))
                self.conexao_db.commit()
                self.logger.info(f"SUCESSO ao processar pedido {numero_pedido}")
                return True
            else:
                # Marca como ERRO e garante que EmailEnviado = 0 para retentar no próximo ciclo
                query = "UPDATE ControleEmailPedidos SET StatusProcessamento = 'ERRO', EmailEnviado = 0, UltimoErro = ?, TentativasEnvio = TentativasEnvio + 1 WHERE Id = ?"
                cursor = self.conexao_db.cursor()
                cursor.execute(query, ('Erro no envio do email', id_controle))
                self.conexao_db.commit()
                self.logger.warning(f"ERRO ao processar pedido {numero_pedido} - Será retentado no próximo ciclo")
                return False
        except Exception as e:
            self.logger.error(f"Erro ao processar pedido {numero_pedido}: {e}")
            # Marca como ERRO e garante que EmailEnviado = 0 para retentar no próximo ciclo
            query = "UPDATE ControleEmailPedidos SET StatusProcessamento = 'ERRO', EmailEnviado = 0, UltimoErro = ?, TentativasEnvio = TentativasEnvio + 1 WHERE Id = ?"
            cursor = self.conexao_db.cursor()
            cursor.execute(query, (str(e)[:500], id_controle))
            self.conexao_db.commit()
            return False

    def validacao_geral(self, pedido):
        """Agrupa todas as validações de um pedido"""
        numero_pedido = pedido['numero']
        enviar_para_cliente = pedido.get('enviar_email_cliente', 1) == 1

        # Se deve enviar para cliente, valida se tem email do cliente
        if enviar_para_cliente and not pedido['email_cliente']:
            self.logger.warning(f"Pedido {numero_pedido}: EnviarEmailCliente=1 mas cliente sem email cadastrado.")
            self.atualizar_status_pedido(pedido['id'], 'StatusProcessamento', 'ERRO_VALIDACAO', 'Cliente sem email')
            return False

        # Se NÃO deve enviar para cliente, valida se tem emails de cópia
        if not enviar_para_cliente and not pedido['emails_copia']:
            self.logger.warning(f"Pedido {numero_pedido}: EnviarEmailCliente=0 mas não há EmailsCopia definidos.")
            self.atualizar_status_pedido(pedido['id'], 'StatusProcessamento', 'ERRO_VALIDACAO', 'EnviarEmailCliente=0 sem EmailsCopia')
            return False

        return True

    def executar_ciclo(self):
        """Executa um ciclo completo de processamento"""
        self.logger.info("=== Iniciando ciclo de processamento ===")
        if not self.conectar_banco(): return
        
        try:
            pedidos = self.buscar_pedidos_para_processar()
            for pedido in pedidos:
                self.processar_pedido(pedido)
        finally:
            if self.conexao_db: self.conexao_db.close()
        self.logger.info("=== Ciclo de processamento concluído ===")

    def testar_deteccao_versoes(self):
        """Função de teste para verificar detecção de versões de PDF sem enviar emails"""
        self.logger.info("=== TESTE: Verificando detecção de versões e sistema de tentativas ===")

        # Listar drivers ODBC disponíveis
        self.logger.info("=== Verificando drivers ODBC instalados ===")
        self.listar_drivers_odbc_disponiveis()

        if not self.conectar_banco():
            self.logger.error("Erro ao conectar ao banco para teste")
            return
        try:
            pedidos = self.buscar_pedidos_para_processar()
            for pedido in pedidos:
                self.logger.info(f"Pedido {pedido['numero']}: Motivo: {pedido['motivo_processamento']}")
        finally:
            if self.conexao_db: self.conexao_db.close()

class ExcelLogger:
    """Classe para gerenciar logs em formato Excel"""
    def __init__(self):
        self.caminho_base = r"C:\Users\Public\Documents\SRPP\scripts"
        os.makedirs(self.caminho_base, exist_ok=True)
        self.arquivo_atual = os.path.join(self.caminho_base, f"log_emails_{datetime.now().strftime('%Y-%m-%d')}.xlsx")

        try:
            if os.path.exists(self.arquivo_atual):
                self.workbook = openpyxl.load_workbook(self.arquivo_atual)
            else:
                self._criar_novo_workbook()
        except Exception as e:
            # Se arquivo estiver corrompido, renomeia e cria novo
            if os.path.exists(self.arquivo_atual):
                backup_name = self.arquivo_atual.replace('.xlsx', f'_corrupted_{datetime.now().strftime("%H%M%S")}.xlsx')
                os.rename(self.arquivo_atual, backup_name)
            self._criar_novo_workbook()

        self.aba_resumo = self.workbook["RESUMO"]
        self.aba_geral = self.workbook["LOG_GERAL"]

    def _criar_novo_workbook(self):
        """Cria um novo workbook Excel"""
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)
        self.workbook.create_sheet("RESUMO")
        self.configurar_aba(self.workbook["RESUMO"], ["Data/Hora", "Pedido", "Cliente", "Email", "Status", "Motivo", "Tentativas", "Versão PDF", "Observações"])
        self.workbook.create_sheet("LOG_GERAL")
        self.configurar_aba(self.workbook["LOG_GERAL"], ["Timestamp", "Pedido", "Cliente", "Fase", "Detalhes", "Validações", "Erro", "Duração", "Thread"])
        self.salvar()

    def configurar_aba(self, aba, cabecalhos):
        """Configura cabeçalhos e formatação de uma aba"""
        aba.append(cabecalhos)
        for cell in aba[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        aba.freeze_panes = "A2"

    def log_resumo(self, **kwargs):
        """Adiciona entrada na aba RESUMO"""
        linha = [datetime.now().strftime("%d/%m/%Y %H:%M:%S")] + list(kwargs.values())
        self.aba_resumo.append(linha)
        self.salvar()

    def log_geral(self, **kwargs):
        """Adiciona entrada na aba LOG_GERAL"""
        linha = [datetime.now().strftime("%d/%m/%Y %H:%M:%S.%f")[:-3]] + list(kwargs.values())
        self.aba_geral.append(linha)
        self.salvar()

    def salvar(self):
        """Salva o arquivo Excel"""
        try:
            self.workbook.save(self.arquivo_atual)
        except PermissionError:
            self.logger.warning("Não foi possível salvar o log do Excel, talvez esteja aberto.")

class PDFEventHandler(FileSystemEventHandler):
    """Handler para monitorar eventos de arquivos PDF"""
    def __init__(self, sistema_emails):
        self.sistema_emails = sistema_emails
        self.aguardar_segundos = int(sistema_emails.get_config('SISTEMA', 'aguardar_segundos_apos_arquivo', fallback=5))
        self.logger = sistema_emails.logger
        
    def on_created(self, event):
        """Chamado quando um arquivo é criado"""
        if not event.is_directory and event.src_path.lower().endswith('.pdf'):
            self.logger.info(f"Novo PDF detectado: {os.path.basename(event.src_path)}")
            time.sleep(self.aguardar_segundos)
            self.sistema_emails.executar_ciclo()
            
    def on_modified(self, event):
        """Chamado quando um arquivo é modificado"""
        if not event.is_directory and event.src_path.lower().endswith('.pdf'):
            self.logger.info(f"PDF modificado detectado: {os.path.basename(event.src_path)}")
            time.sleep(self.aguardar_segundos)
            self.sistema_emails.executar_ciclo()

def main():
    """Função principal"""
    def signal_handler(sig, frame):
        print('\nEncerrando sistema...')
        sys.exit(0)
    
    signal.signal(signal.SIGINT, signal_handler)
    
    try:
        sistema = SistemaEnvioEmails()
        if len(sys.argv) > 1 and sys.argv[1] == "--teste":
            print("MODO TESTE: Verificando detecção de versões (SEM ENVIAR EMAILS)")
            sistema.testar_deteccao_versoes()
            return

        print("Sistema de Envio de Emails iniciado!")
        caminho_pdfs = sistema.get_config('PDFS', 'caminho')
        print(f"Monitorando pasta: {caminho_pdfs}")

        verificacao_inicial = sistema.get_config('SISTEMA', 'verificacao_inicial', fallback=True)
        if isinstance(verificacao_inicial, str):
            verificacao_inicial = verificacao_inicial.lower() in ('true', '1', 'yes')
        if verificacao_inicial:
            print("Executando verificação inicial...")
            sistema.executar_ciclo()

        event_handler = PDFEventHandler(sistema)
        observer = Observer()
        observer.schedule(event_handler, caminho_pdfs, recursive=False)
        observer.start()
        
        print("Pressione Ctrl+C para encerrar.")
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            observer.stop()
        observer.join()
        print("Sistema encerrado.")
    except Exception as e:
        print(f"Erro fatal ao inicializar: {e}")

if __name__ == "__main__":
    main()